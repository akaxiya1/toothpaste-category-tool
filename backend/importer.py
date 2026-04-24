from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config import IMPORT_DIR, TEMP_DIR, ensure_directories
from .constants import FIELD_SET_BY_KIND
from .logic import enrich_candidate, enrich_sku, normalize_header, normalize_text, similarity


def save_upload(file_name: str, content: bytes) -> Path:
    ensure_directories()
    safe_name = Path(file_name).name or "upload.bin"
    target = TEMP_DIR / safe_name
    counter = 1
    while target.exists():
        stem = Path(safe_name).stem
        suffix = Path(safe_name).suffix
        target = TEMP_DIR / f"{stem}_{counter}{suffix}"
        counter += 1
    target.write_bytes(content)
    return target


def persist_import_copy(source_path: Path) -> Path:
    ensure_directories()
    target = IMPORT_DIR / source_path.name
    counter = 1
    while target.exists():
        target = IMPORT_DIR / f"{source_path.stem}_{counter}{source_path.suffix}"
        counter += 1
    target.write_bytes(source_path.read_bytes())
    return target


def load_tabular_rows(file_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        return _load_xlsx(file_path)
    return _load_csv(file_path)


def _load_xlsx(file_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [normalize_text(value) or f"列{i+1}" for i, value in enumerate(rows[0])]
    items: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(value not in {None, ""} for value in row):
            continue
        items.append({headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))})
    return headers, items


def _load_csv(file_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    encodings = ["utf-8-sig", "gb18030", "utf-16"]
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            text = file_path.read_text(encoding=encoding)
            reader = csv.DictReader(io.StringIO(text))
            headers = reader.fieldnames or []
            rows = [row for row in reader if any(normalize_text(value) for value in row.values())]
            return headers, rows
        except Exception as exc:  # pragma: no cover - fallback branch
            last_error = exc
    raise ValueError(f"无法解析文件编码: {file_path.name}") from last_error


def suggest_mapping(kind: str, headers: list[str]) -> dict[str, str]:
    fields = FIELD_SET_BY_KIND[kind]
    normalized_headers = {normalize_header(header): header for header in headers}
    mapping: dict[str, str] = {}
    for field in fields:
        wanted = [field["label"], *field.get("aliases", [])]
        selected = ""
        for candidate in wanted:
            match = normalized_headers.get(normalize_header(candidate))
            if match:
                selected = match
                break
        if not selected:
            for header in headers:
                if similarity(normalize_header(header), normalize_header(field["label"])) >= 0.75:
                    selected = header
                    break
        mapping[field["key"]] = selected
    return mapping


def missing_required_fields(kind: str, mapping: dict[str, str]) -> list[str]:
    return [field["label"] for field in FIELD_SET_BY_KIND[kind] if field["required"] and not mapping.get(field["key"])]


def preview_import(kind: str, file_path: Path, existing_skus: list[dict[str, Any]]) -> dict[str, Any]:
    headers, rows = load_tabular_rows(file_path)
    mapping = suggest_mapping(kind, headers)
    missing = missing_required_fields(kind, mapping)
    sample_rows = rows[:5]
    exact_duplicates: list[dict[str, Any]] = []
    similar_duplicates: list[dict[str, Any]] = []

    if kind == "sku":
        catalog_by_code = {normalize_text(item.get("sku_code")): item for item in existing_skus}
        seen_codes: set[str] = set()
        for row in rows[:100]:
            sku_code_column = mapping.get("sku_code")
            name_column = mapping.get("product_name")
            row_code = normalize_text(row.get(sku_code_column)) if sku_code_column else ""
            row_name = normalize_text(row.get(name_column)) if name_column else ""
            if row_code and row_code in seen_codes:
                exact_duplicates.append({"sku_code": row_code, "existing_name": "本次导入文件内重复"})
            if row_code:
                seen_codes.add(row_code)
            if row_code and row_code in catalog_by_code:
                exact_duplicates.append({"sku_code": row_code, "existing_name": catalog_by_code[row_code]["product_name"]})
            if row_name:
                for item in existing_skus:
                    if similarity(row_name, normalize_text(item.get("product_name"))) >= 0.86 and row_code != normalize_text(item.get("sku_code")):
                        similar_duplicates.append(
                            {
                                "incoming_name": row_name,
                                "existing_name": item["product_name"],
                                "existing_sku": item["sku_code"],
                            }
                        )
                        break
    return {
        "headers": headers,
        "row_count": len(rows),
        "sample_rows": sample_rows,
        "mapping": mapping,
        "missing_required": missing,
        "duplicate_skus": exact_duplicates[:20],
        "similar_names": similar_duplicates[:20],
    }


def commit_import(kind: str, file_path: Path, mapping: dict[str, str], existing_skus: list[dict[str, Any]]) -> list[dict[str, Any]]:
    _, rows = load_tabular_rows(file_path)
    items: list[dict[str, Any]] = []
    for row in rows:
        payload = {key: row.get(source_column, "") for key, source_column in mapping.items() if source_column}
        if not any(normalize_text(value) for value in payload.values()):
            continue
        if kind == "sku":
            enriched = enrich_sku(payload)
            if not enriched["sku_code"] or not enriched["brand"] or not enriched["product_name"]:
                continue
            items.append(enriched)
        else:
            enriched = enrich_candidate(payload, existing_skus)
            if not enriched["brand"] or not enriched["product_name"]:
                continue
            items.append(enriched)
    return items
